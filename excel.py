import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename, new_filename):
    wb = xl.load_workbook(filename)
    sheet = wb['上BI']

    cell = sheet.cell(1, 1)

    # cell = sheet['B3']

    print("Cell A1:"+cell.value)

    """
    # How to add BarChart 
    
    values = Reference(sheet,
              min_row=3,
              max_row=sheet.max_row,
              min_col=2,
              max_col=2)

    chart = BarChart()
    chart.add_data(values)
    coordinate = f'A{sheet.max_row+2}'
    print(coordinate)

    #chart_cell = sheet.cell(sheet.max_row, 1)

    sheet.add_chart(chart, coordinate)
    """

    wb.save(f'{new_filename}')


process_workbook("BI任务汇总表(1).xlsx", "BI任务汇总表(1)_copy.xlsx")